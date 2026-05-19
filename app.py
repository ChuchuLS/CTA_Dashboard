"""
Pulsar — Global Rates & Equity Scoring Dashboards
==================================================
Streamlit app replicating the Pulsar visual style for two dashboards:
- Global Rates Scoring  (10 sovereign bond markets)
- Global Equity Scoring (17 equity index futures)

Run:
    streamlit run app.py

Repo layout:
    .
    ├── app.py
    ├── requirements.txt
    ├── README.md
    └── data/
        └── pulsar_data.xlsx
"""

from pathlib import Path
import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

APP_DIR = Path(__file__).parent
DATA_PATH = APP_DIR / "data" / "pulsar_data.xlsx"


# ============================================================
# UNIVERSES & MAPPINGS
# ============================================================

# Rates dashboard: 10 sovereign bond markets
RATES_UNIVERSE = [
    ("FR", "France"),
    ("JP", "Japan"),
    ("GB", "UK"),
    ("ES", "Spain"),
    ("IT", "Italy"),
    ("US", "United States"),
    ("CA", "Canada"),
    ("KR", "S. Korea"),
    ("AU", "Australia"),
    ("DE", "Germany"),
]
RATES_CODES = [c for c, _ in RATES_UNIVERSE]

# Equity dashboard: 17 indices
EQUITY_UNIVERSE = [
    ("PT1",  "S&P/TSX",        "Canada"),
    ("NQ1",  "Nasdaq 100",     "USA"),
    ("KM1",  "KOSPI",          "S. Korea"),
    ("XP1",  "ASX",            "Australia"),
    ("HI1",  "Hang Seng",      "Hong Kong"),
    ("XU1",  "FTSE China A50", "China"),
    ("ES1",  "S&P 500",        "USA"),
    ("RTY1", "Russell 2000",   "USA"),
    ("SM1",  "SMI",            "Switzerland"),
    ("Z 1",  "FTSE 100",       "UK"),
    ("EO1",  "AEX",            "Netherlands"),
    ("CF1",  "CAC 40",         "France"),
    ("ST1",  "FTSE MIB",       "Italy"),
    ("NK1",  "Nikkei 225",     "Japan"),
    ("VG1",  "Euro Stoxx 50",  "Eurozone"),
    ("IB1",  "IBEX 35",        "Spain"),
    ("GX1",  "DAX",            "Germany"),
]
EQUITY_CODES = [c for c, _, _ in EQUITY_UNIVERSE]
EQUITY_META = {c: (n, r) for c, n, r in EQUITY_UNIVERSE}

# Each equity index → its macro country (for GDP/CPI/Fiscal lookup)
INDEX_TO_COUNTRY = {
    "PT1": "CA", "NQ1": "US", "KM1": "KR", "XP1": "AU",
    "HI1": "HK", "XU1": "CN", "ES1": "US", "RTY1": "US",
    "SM1": "CH", "Z 1": "GB", "EO1": "NL", "CF1": "FR",
    "ST1": "IT", "NK1": "JP", "VG1": "EZ", "IB1": "ES", "GX1": "DE",
}

# Each equity index → its FCI region (we have 4 FCIs only)
# Goldman/Bloomberg only publish FCI for major financial systems
INDEX_TO_FCI_REGION = {
    "PT1": "NQ1",   # Canada → US FCI (closest financial system proxy)
    "NQ1": "NQ1",   # US
    "KM1": "NQ1",   # Korea → US FCI (USD-funded EM)
    "XP1": "NQ1",   # Australia → US FCI
    "HI1": "XU1",   # Hang Seng → China FCI
    "XU1": "XU1",   # China
    "ES1": "NQ1",   # US
    "RTY1": "NQ1",  # US
    "SM1": "VG1",   # Switzerland → Eurozone FCI
    "Z 1": "Z 1",   # UK
    "EO1": "VG1",   # Netherlands → Eurozone FCI
    "CF1": "VG1",   # France → Eurozone FCI
    "ST1": "VG1",   # Italy → Eurozone FCI
    "NK1": "NQ1",   # Japan → US FCI (open economy fallback)
    "VG1": "VG1",   # Eurozone
    "IB1": "VG1",   # Spain → Eurozone FCI
    "GX1": "VG1",   # Germany → Eurozone FCI
}

# Each equity index → its Citi ToT currency ticker
INDEX_TO_TOT_TICKER = {
    "PT1": "CTOTCAD Index", "NQ1": "CTOTUSD Index", "KM1": "CTOTKRW Index",
    "XP1": "CTOTAUD Index", "HI1": "CTOTHKD Index", "XU1": "CTOTCNY Index",
    "ES1": "CTOTUSD Index", "RTY1": "CTOTUSD Index", "SM1": "CTOTCHF Index",
    "Z 1": "CTOTGBP Index", "EO1": "CTOTEUR Index", "CF1": "CTOTEUR Index",
    "ST1": "CTOTEUR Index", "NK1": "CTOTJPY Index", "VG1": "CTOTEUR Index",
    "IB1": "CTOTEUR Index", "GX1": "CTOTEUR Index",
}

# Each rates country → its Citi ToT currency
RATES_TO_TOT_TICKER = {
    "FR": "CTOTEUR Index", "JP": "CTOTJPY Index", "GB": "CTOTGBP Index",
    "ES": "CTOTEUR Index", "IT": "CTOTEUR Index", "US": "CTOTUSD Index",
    "CA": "CTOTCAD Index", "KR": "CTOTKRW Index", "AU": "CTOTAUD Index",
    "DE": "CTOTEUR Index",
}


# ============================================================
# LOADER
# ============================================================
@st.cache_data(show_spinner=False)
def read_sheet(xlsx_path: str, sheet_name: str, header_row: int = 4) -> pd.DataFrame:
    """
    Read a sheet with row 4 as code headers, dates in col A from row 6.
    Handles BDH-style 'Date' interleaved columns by skipping them.
    Returns DataFrame indexed by date with code columns.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    # Read row 4 codes; drop any column literally labeled "Date"
    raw_codes = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        raw_codes.append((c, str(v).strip() if v is not None else None))

    # Keep only columns whose code is not None and not "Date"
    valid_cols = [(c, code) for c, code in raw_codes if code and code != "Date"]
    codes = [code for _, code in valid_cols]
    col_idx = [c for c, _ in valid_cols]

    rows = []
    for r in range(6, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if d is None or (isinstance(d, str) and (d.strip() == "" or "#N/A" in d)):
            continue
        try:
            ts = pd.Timestamp(d)
        except Exception:
            continue
        vals = []
        for c in col_idx:
            v = ws.cell(row=r, column=c).value
            if v is None or v == "" or (isinstance(v, str) and "#N/A" in v):
                vals.append(np.nan)
            else:
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    vals.append(np.nan)
        rows.append([ts] + vals)

    df = pd.DataFrame(rows, columns=["date"] + codes).set_index("date").sort_index()
    df.index = pd.to_datetime(df.index)
    df = df[~df.index.duplicated(keep="last")]
    return df


@st.cache_data(show_spinner=False)
def load_all(xlsx_path: str) -> dict:
    return {
        "gdp":    read_sheet(xlsx_path, "Macro_GDP"),
        "cpi":    read_sheet(xlsx_path, "Macro_CPI"),
        "fiscal": read_sheet(xlsx_path, "Macro_Fiscal"),
        "y10y":   read_sheet(xlsx_path, "Rates_10Y"),
        "tot":    read_sheet(xlsx_path, "Equity_ToT"),
        "fci":    read_sheet(xlsx_path, "Equity_FCI"),
        "eps":    read_sheet(xlsx_path, "Equity_EPS"),
        "px":     read_sheet(xlsx_path, "Equity_Prices"),
    }


# ============================================================
# FACTOR MATH HELPERS
# ============================================================
def latest(df: pd.DataFrame, asof: pd.Timestamp) -> pd.Series:
    """Latest available value per column on or before asof."""
    sub = df.loc[:asof].ffill()
    if len(sub) == 0:
        return pd.Series(dtype=float)
    return sub.iloc[-1]


def value_n_days_ago(df: pd.DataFrame, asof: pd.Timestamp, days: int) -> pd.Series:
    sub = df.loc[:asof].ffill()
    if len(sub) == 0:
        return pd.Series(dtype=float)
    target = asof - pd.Timedelta(days=days)
    if len(sub.loc[:target]) > 0:
        return sub.loc[:target].iloc[-1]
    return sub.iloc[0]


def pct_change(df: pd.DataFrame, asof: pd.Timestamp, days: int) -> pd.Series:
    last = latest(df, asof)
    base = value_n_days_ago(df, asof, days)
    return (last / base - 1) * 100


def diff_n_days(df: pd.DataFrame, asof: pd.Timestamp, days: int) -> pd.Series:
    """Absolute (not %) change over n days. For yields, ToT etc."""
    last = latest(df, asof)
    base = value_n_days_ago(df, asof, days)
    return last - base


def realized_vol(df: pd.DataFrame, asof: pd.Timestamp, window: int = 30) -> pd.Series:
    """Annualized rolling std of daily log returns."""
    sub = df.loc[:asof].ffill()
    if len(sub) < 2:
        return pd.Series(index=df.columns, dtype=float)
    rets = np.log(sub).diff()
    tail = rets.tail(window)
    return tail.std() * np.sqrt(252) * 100


def zscore(series: pd.Series, codes: list, sign: int = 1) -> pd.Series:
    """Cross-sectional z-score over the given universe; sign flips direction."""
    s = series.reindex(codes).astype(float)
    mu = s.mean(skipna=True)
    sd = s.std(ddof=0, skipna=True)
    if sd == 0 or np.isnan(sd):
        return s * 0
    return sign * (s - mu) / sd


# ============================================================
# RATES SCORING
# ============================================================
def score_rates(data: dict, asof: pd.Timestamp, weights: dict) -> pd.DataFrame:
    """
    Replicate the Pulsar Rates dashboard:
    Macro pillar = mean(GDP_z, CPI_z (inverted), Budget_z)
    Markets pillar = mean(Momentum_z, Carry_z, RealYield_z)
    Score = mean(Macro, Markets)
    """
    gdp_v    = latest(data["gdp"],    asof)
    cpi_v    = latest(data["cpi"],    asof)
    fiscal_v = latest(data["fiscal"], asof)
    y10y_v   = latest(data["y10y"],   asof)
    y10y_3m  = value_n_days_ago(data["y10y"], asof, 90)

    # Build factor frames per country (only rates universe)
    z_gdp    = zscore(gdp_v,    RATES_CODES, +1)   # higher GDP = good for govt
    z_cpi    = zscore(cpi_v,    RATES_CODES, -1)   # lower inflation = good for bonds
    z_budget = zscore(fiscal_v, RATES_CODES, +1)   # less deficit (less negative) = good
    macro    = pd.concat([z_gdp, z_cpi, z_budget], axis=1).mean(axis=1)

    # 3M yield change: bonds rally when yields FALL → lower change = positive
    z_mom    = zscore(y10y_v - y10y_3m, RATES_CODES, -1)
    z_carry  = zscore(y10y_v,            RATES_CODES, +1)   # higher carry = good
    # Real yield = nominal − CPI
    real_y   = y10y_v - cpi_v
    z_realy  = zscore(real_y,            RATES_CODES, +1)
    markets  = pd.concat([z_mom, z_carry, z_realy], axis=1).mean(axis=1)

    # Weighted composite
    wsum = sum(weights.values())
    wn = {k: v / wsum for k, v in weights.items()}
    score = macro * wn["macro"] + markets * wn["markets"]

    out = pd.DataFrame({
        "gdp_z":    z_gdp,
        "cpi_z":    z_cpi,
        "budget_z": z_budget,
        "macro":    macro,
        "mom_z":    z_mom,
        "carry_z":  z_carry,
        "realy_z":  z_realy,
        "markets":  markets,
        "score":    score,
    }).round(2)
    out.index.name = "code"
    out["country"] = [dict(RATES_UNIVERSE)[c] for c in out.index]
    out["incomplete"] = pd.concat([macro, markets], axis=1).isna().any(axis=1)
    return out.sort_values("score", ascending=False, na_position="last")


# ============================================================
# EQUITY SCORING
# ============================================================
def score_equity(data: dict, asof: pd.Timestamp, weights: dict) -> pd.DataFrame:
    """
    Pulsar Equity dashboard:
    Macro pillar = mean(Growth_z, Inflation_z (inv), Deficit_z, ToT_z, FCI_z)
    Then composite score = weighted blend of macro + EPS revisions.
    Plus performance and vol columns.
    """
    gdp_v    = latest(data["gdp"],    asof)
    cpi_v    = latest(data["cpi"],    asof)
    fiscal_v = latest(data["fiscal"], asof)
    tot_v    = latest(data["tot"],    asof)
    tot_3m   = value_n_days_ago(data["tot"], asof, 90)
    fci_v    = latest(data["fci"],    asof)

    # Map each index to its country macro values
    def by_index(country_series: pd.Series) -> pd.Series:
        out = {}
        for code in EQUITY_CODES:
            country = INDEX_TO_COUNTRY[code]
            out[code] = country_series.get(country, np.nan)
        return pd.Series(out)

    growth     = by_index(gdp_v)
    inflation  = by_index(cpi_v)
    deficit    = by_index(fiscal_v)

    # ToT: 3M change in the index's currency ToT
    def tot_for_index(asof_series, three_m_series):
        out_now, out_then = {}, {}
        for code in EQUITY_CODES:
            t = INDEX_TO_TOT_TICKER.get(code)
            out_now[code]  = asof_series.get(t, np.nan)
            out_then[code] = three_m_series.get(t, np.nan)
        return pd.Series(out_now), pd.Series(out_then)
    tot_now, tot_then = tot_for_index(tot_v, tot_3m)
    tot_mom = tot_now - tot_then  # absolute change in ToT index

    # FCI: latest value of the index's regional FCI
    fci_by_index = pd.Series({code: fci_v.get(INDEX_TO_FCI_REGION[code], np.nan)
                              for code in EQUITY_CODES})

    # EPS Δ: 3M % change in FY1 EPS estimate
    eps_v   = latest(data["eps"], asof)
    eps_3m  = value_n_days_ago(data["eps"], asof, 90)
    eps_chg = (eps_v / eps_3m - 1) * 100
    # Align to equity universe (eps frame columns are already index codes)
    eps_delta = eps_chg.reindex(EQUITY_CODES)

    # Z-scores on equity universe
    z_growth = zscore(growth,    EQUITY_CODES, +1)
    z_infl   = zscore(inflation, EQUITY_CODES, -1)
    z_def    = zscore(deficit,   EQUITY_CODES, +1)
    z_tot    = zscore(tot_mom,   EQUITY_CODES, +1)
    z_fci    = zscore(fci_by_index, EQUITY_CODES, +1)

    macro = pd.concat([z_growth, z_infl, z_def, z_tot, z_fci], axis=1).mean(axis=1)

    # EPS as separate factor (z-scored across the panel)
    z_eps = zscore(eps_delta, EQUITY_CODES, +1)

    # Weighted composite
    wsum = sum(weights.values())
    wn = {k: v / wsum for k, v in weights.items()}
    score = macro * wn["macro"] + z_eps * wn["eps"]

    # Performance and vol (cosmetic columns)
    p5d = pct_change(data["px"], asof, 7).reindex(EQUITY_CODES)
    p1m = pct_change(data["px"], asof, 30).reindex(EQUITY_CODES)
    p3m = pct_change(data["px"], asof, 90).reindex(EQUITY_CODES)
    vol = realized_vol(data["px"], asof, 30).reindex(EQUITY_CODES)

    out = pd.DataFrame({
        "growth_z": z_growth,
        "infl_z":   z_infl,
        "def_z":    z_def,
        "tot_z":    z_tot,
        "fci_z":    z_fci,
        "macro":    macro,
        "eps_delta": eps_delta,
        "score":    score,
        "p5d": p5d,
        "p1m": p1m,
        "p3m": p3m,
        "vol": vol,
    }).round(2)
    out.index.name = "code"
    out["name"]   = [EQUITY_META[c][0] for c in out.index]
    out["region"] = [EQUITY_META[c][1] for c in out.index]
    out["incomplete"] = macro.isna() | z_eps.isna()
    return out.sort_values("score", ascending=False, na_position="last")


# ============================================================
# RENDERING (Pulsar HTML style)
# ============================================================
PULSAR_CSS = """
:root {
  --paper: #ece7d5;
  --ink: #15140f;
  --ink-soft: #2a271e;
  --muted: #8a8470;
  --rule: #15140f;
  --green: #2f5b2a;
  --green-deep: #1f3f1c;
  --red: #8a2a1c;
  --red-deep: #5a1a10;
  --orange: #d28a1f;
  --shade: rgba(180, 165, 120, 0.18);
}
* { box-sizing: border-box; margin: 0; padding: 0; }
html, body {
  background: var(--paper);
  color: var(--ink);
  font-family: 'JetBrains Mono', monospace;
  font-size: 13px;
  -webkit-font-smoothing: antialiased;
}
body { padding: 28px 40px 36px; }
.masthead {
  display: flex; justify-content: space-between; align-items: baseline;
  font-size: 12px; letter-spacing: 0.04em; padding-bottom: 8px;
}
.masthead-left { display: flex; align-items: center; gap: 16px; }
.brand-dot {
  width: 9px; height: 9px; border-radius: 50%;
  background: var(--orange); display: inline-block;
  margin-right: 6px; transform: translateY(1px);
}
.brand { font-weight: 700; letter-spacing: 0.06em; }
.week, .masthead-right { color: var(--ink-soft); letter-spacing: 0.06em; }
.double-rule {
  border-top: 1px solid var(--rule);
  border-bottom: 1px solid var(--rule);
  height: 4px; margin-bottom: 28px;
}
.title-row {
  display: grid; grid-template-columns: 1fr 340px;
  align-items: end; gap: 32px; margin-bottom: 28px;
}
h1 {
  font-family: 'PT Serif', serif; font-style: italic; font-weight: 700;
  font-size: 64px; line-height: 0.98; letter-spacing: -0.02em;
}
.blurb {
  font-size: 11.5px; letter-spacing: 0.04em; line-height: 1.55;
  text-align: right; color: var(--ink-soft); text-transform: uppercase;
}
.table { width: 100%; border-collapse: collapse; }
.table thead th {
  font-weight: 500; font-size: 10.5px; letter-spacing: 0.06em;
  text-align: right; padding: 0 6px 14px;
  color: var(--ink); vertical-align: bottom; line-height: 1.4;
}
.table thead th.col-rank,
.table thead th.col-flag,
.table thead th.col-name { text-align: left; }
.table thead th .sub {
  display: block; font-size: 9.5px; color: var(--muted);
  font-weight: 400; letter-spacing: 0.06em; margin-top: 4px;
}
.group-head { color: var(--orange); }
.table tbody tr { border-top: 1px solid rgba(21, 20, 15, 0.18); }
.table tbody tr:last-child { border-bottom: 1px solid var(--rule); }
.table tbody td {
  padding: 14px 6px; font-size: 12.5px; text-align: right;
  vertical-align: middle; font-weight: 500;
}
.col-rank { text-align: left !important; width: 56px; padding-left: 0 !important; }
.col-flag { text-align: left !important; width: 32px; color: var(--muted); font-size: 10px; letter-spacing: 0.06em; }
.col-name { text-align: left !important; width: 170px; }
.col-name .name {
  font-family: 'PT Serif', serif; font-style: italic; font-weight: 700;
  font-size: 18px; letter-spacing: -0.005em; color: var(--ink);
  display: block; line-height: 1.15;
}
.col-name .sub {
  display: block; margin-top: 2px; color: var(--muted);
  font-size: 10px; letter-spacing: 0.06em; text-transform: uppercase;
}
.col-score { font-weight: 700; width: 64px; padding-right: 0 !important; }
.shade { background: var(--shade); }
.pip {
  display: inline-flex; align-items: center; justify-content: center;
  width: 26px; height: 26px; border-radius: 50%;
  font-size: 12px; font-weight: 700; color: var(--paper);
}
.pip.plain { background: none; color: var(--ink); font-weight: 500; }
.pip.green { background: var(--green-deep); }
.pip.orange { background: var(--orange); }
.pip.red { background: var(--red-deep); }
.pos { color: var(--green); }
.neg { color: var(--red); }
.neutral { color: var(--ink-soft); }
.footer {
  margin-top: 28px;
  display: grid; grid-template-columns: 1fr 1fr 1fr;
  gap: 40px; font-size: 11px; line-height: 1.6; color: var(--ink-soft);
}
.footer h4 {
  font-size: 11px; font-weight: 500; letter-spacing: 0.06em;
  color: var(--orange); margin-bottom: 10px; text-transform: uppercase;
}
.footer h4::before { content: "— "; }
"""


def fmt_num(v, dp=2, pct=False):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ("neutral", "n/a")
    sign = "+" if v >= 0 else ""
    s = f"{sign}{v:.{dp}f}{'%' if pct else ''}"
    cls = "pos" if v > 0 else ("neg" if v < 0 else "neutral")
    return (cls, s)


def pip_class(rank, n_total, incomplete, score, pillars):
    if incomplete:
        return "orange"
    if score is not None and not (isinstance(score, float) and np.isnan(score)):
        for p in pillars:
            if p is not None and not (isinstance(p, float) and np.isnan(p)):
                if abs(p) > 1.0 and (p > 0) != (score > 0):
                    return "orange"
    if rank == 1 or rank == 2:
        return "green"
    if rank == n_total:
        return "red"
    return "plain"


# --- Rates dashboard HTML ---
COUNTRY_ISO = {
    "France": "FR", "Japan": "JP", "UK": "GB", "Spain": "ES", "Italy": "IT",
    "United States": "US", "Canada": "CA", "S. Korea": "KR",
    "Australia": "AU", "Germany": "DE",
}

def render_rates(scores: pd.DataFrame, asof: str, weights: dict) -> str:
    n = len(scores)
    tbody = []
    for i, (code, row) in enumerate(scores.iterrows(), start=1):
        country = row["country"]
        iso = COUNTRY_ISO.get(country, code)
        pip = pip_class(i, n, row["incomplete"], row["score"],
                        [row["macro"], row["markets"]])
        def td(v, shade=False):
            cls, s = fmt_num(v)
            return f'<td class="{cls}{" shade" if shade else ""}">{s}</td>'
        score_cls, score_s = fmt_num(row["score"])
        tbody.append(f"""
          <tr>
            <td class="col-rank"><span class="pip {pip}">{i}</span></td>
            <td class="col-flag">{iso}</td>
            <td class="col-name"><span class="name">{country}</span></td>
            {td(row["gdp_z"])}
            {td(row["cpi_z"])}
            {td(row["budget_z"])}
            {td(row["macro"], shade=True)}
            {td(row["mom_z"])}
            {td(row["carry_z"])}
            {td(row["realy_z"])}
            {td(row["markets"], shade=True)}
            <td class="col-score {score_cls}">{score_s}</td>
          </tr>""")

    w_str = f"Macro {weights['macro']:.0f} · Markets {weights['markets']:.0f}"
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=PT+Serif:ital,wght@0,400;0,700;1,400;1,700&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>{PULSAR_CSS}</style></head><body>
<header class="masthead">
  <div class="masthead-left">
    <span><span class="brand-dot"></span><span class="brand">PULSAR</span></span>
    <span class="week">/ RATES</span>
  </div>
  <div class="masthead-right">{asof.upper()} · CROSS-ASSET</div>
</header>
<div class="double-rule"></div>
<div class="title-row">
  <h1>Global Rates<br>Scoring</h1>
  <p class="blurb">COMPOSITE SCORES ACROSS 10<br>SOVEREIGN BOND MARKETS, BLENDING<br>MACRO &amp; MARKET FACTORS.<br>WEIGHTS: {w_str.upper()}</p>
</div>
<table class="table"><thead><tr>
  <th class="col-rank">RANK</th><th class="col-flag"></th><th class="col-name">COUNTRY</th>
  <th class="group-head">MACRO<span class="sub">GDP</span></th><th>CPI</th><th>BUDGET</th>
  <th class="shade">MACRO</th>
  <th class="group-head">MARKET<span class="sub">MOM.</span></th><th>CARRY</th><th>REAL Y.</th>
  <th class="shade">MARKETS</th><th class="col-score">SCORE</th>
</tr></thead><tbody>{''.join(tbody)}</tbody></table>
<footer class="footer">
  <div><h4>METHODOLOGY</h4><p>Composite z-scores across macro fundamentals (GDP, CPI inverted, fiscal balance) and market signals (3M yield momentum inverted, 10Y carry, real yield). Pillars weighted per sidebar. As of {asof}.</p></div>
  <div><h4>READING THE HEAT</h4><p>Green tones indicate constructive readings for the bond market; red tones flag deterioration. Pillar composites shaded. Orange pip flags pillar-score tension or incomplete data.</p></div>
  <div><h4>SOURCE</h4><p>Bloomberg · Internal model<br>Data as of close, prior session</p></div>
</footer></body></html>"""


# --- Equity dashboard HTML ---
def render_equity(scores: pd.DataFrame, asof: str, weights: dict) -> str:
    n = len(scores)
    tbody = []
    for i, (code, row) in enumerate(scores.iterrows(), start=1):
        pip = pip_class(i, n, row["incomplete"], row["score"], [row["macro"]])
        def td(v, dp=2, pct=False, shade=False):
            cls, s = fmt_num(v, dp, pct)
            return f'<td class="{cls}{" shade" if shade else ""}">{s}</td>'
        score_cls, score_s = fmt_num(row["score"])
        vol_val = row["vol"]
        vol_s = "n/a" if pd.isna(vol_val) else f"{vol_val:.2f}%"
        tbody.append(f"""
          <tr>
            <td class="col-rank"><span class="pip {pip}">{i}</span></td>
            <td class="col-name"><span class="name">{row["name"]}</span><span class="sub">{code} · {row["region"]}</span></td>
            {td(row["growth_z"])}
            {td(row["infl_z"])}
            {td(row["def_z"])}
            {td(row["tot_z"])}
            {td(row["fci_z"])}
            {td(row["macro"], shade=True)}
            {td(row["eps_delta"], pct=True)}
            <td class="col-score {score_cls}">{score_s}</td>
            {td(row["p5d"], pct=True)}
            {td(row["p1m"], pct=True)}
            {td(row["p3m"], pct=True)}
            <td class="neutral">{vol_s}</td>
          </tr>""")

    w_str = f"Macro {weights['macro']:.0f} · EPS {weights['eps']:.0f}"
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=PT+Serif:ital,wght@0,400;0,700;1,400;1,700&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>{PULSAR_CSS}</style></head><body>
<header class="masthead">
  <div class="masthead-left">
    <span><span class="brand-dot"></span><span class="brand">PULSAR</span></span>
    <span class="week">/ EQUITY</span>
  </div>
  <div class="masthead-right">{asof.upper()} · CROSS-ASSET</div>
</header>
<div class="double-rule"></div>
<div class="title-row">
  <h1>Global Equity<br>Scoring</h1>
  <p class="blurb">COMPOSITE SCORES ACROSS 17 EQUITY<br>INDEX FUTURES, BLENDING MACRO,<br>EPS REVISIONS &amp; TRAILING PERF.<br>WEIGHTS: {w_str.upper()}</p>
</div>
<table class="table"><thead><tr>
  <th class="col-rank">RANK</th><th class="col-name">INDEX</th>
  <th class="group-head">MACRO<span class="sub">GROWTH</span></th>
  <th>INFL.</th><th>DEFICIT</th><th>TOT</th><th>FCI</th>
  <th class="shade">MACRO</th>
  <th class="group-head">EPS Δ<span class="sub">3M %</span></th>
  <th class="col-score">SCORE</th>
  <th class="group-head">PERF.<span class="sub">5D</span></th><th>1M</th><th>3M</th><th>VOL</th>
</tr></thead><tbody>{''.join(tbody)}</tbody></table>
<footer class="footer">
  <div><h4>METHODOLOGY</h4><p>Composite z-scores across 5 macro factors (growth, inflation inv., deficit, terms of trade 3M, FCI) and bottom-up EPS revisions (3M % change in FY1). Equity macro is keyed by country via index→country map; FCI is keyed by region (US/EZ/UK/CN). Weights configurable. As of {asof}.</p></div>
  <div><h4>READING THE HEAT</h4><p>Green tones flag constructive readings; red tones flag deterioration. Macro composite shaded. Orange pip flags pillar-score tension or incomplete data coverage.</p></div>
  <div><h4>SOURCE</h4><p>Bloomberg · Internal model<br>Data as of close, prior session</p></div>
</footer></body></html>"""


# ============================================================
# STREAMLIT APP
# ============================================================
st.set_page_config(
    page_title="Pulsar — Cross-Asset Scoring",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.stApp { background: #ece7d5; }
section[data-testid="stSidebar"] { background: #e3deca; }
section[data-testid="stSidebar"] * { color: #15140f !important; }
h1, h2, h3, .stMarkdown { color: #15140f; }
.block-container { padding-top: 1.5rem; padding-bottom: 1rem; max-width: 100%; }
</style>
""", unsafe_allow_html=True)

st.sidebar.markdown("### Pulsar")
st.sidebar.caption("Cross-asset scoring dashboards")

if not DATA_PATH.exists():
    st.sidebar.error(f"Data file not found: {DATA_PATH}")
    st.stop()

try:
    data = load_all(str(DATA_PATH))
except Exception as e:
    st.error(f"Failed to load workbook: {e}")
    st.stop()

# Dashboard picker
dash = st.sidebar.radio("Dashboard", ["Rates", "Equity"], horizontal=True)

# As-of date — bounded by the most-restrictive daily data range
px_max = data["px"].index.max() if len(data["px"]) else pd.Timestamp.today()
y10_max = data["y10y"].index.max() if len(data["y10y"]) else pd.Timestamp.today()
asof_max = min(px_max, y10_max) if dash == "Equity" else y10_max
asof_min = asof_max - pd.Timedelta(days=90)
asof = st.sidebar.date_input(
    "As-of date", value=asof_max.date(),
    min_value=asof_min.date(), max_value=asof_max.date(),
)
asof = pd.Timestamp(asof)

st.sidebar.markdown("---")
st.sidebar.markdown("**Pillar weights**")
if dash == "Rates":
    w_macro = st.sidebar.slider("Macro",   0.0, 1.0, 0.50, 0.01)
    w_mkt   = st.sidebar.slider("Markets", 0.0, 1.0, 0.50, 0.01)
    weights = {"macro": w_macro, "markets": w_mkt}
else:
    w_macro = st.sidebar.slider("Macro", 0.0, 1.0, 0.50, 0.01)
    w_eps   = st.sidebar.slider("EPS Δ", 0.0, 1.0, 0.50, 0.01)
    weights = {"macro": w_macro, "eps": w_eps}

if sum(weights.values()) == 0:
    st.sidebar.error("Weights cannot all be zero.")
    st.stop()

st.sidebar.markdown("---")
show_raw  = st.sidebar.checkbox("Show raw factor table", value=False)
show_diag = st.sidebar.checkbox("Show data diagnostics", value=False)

# Compute and render
asof_str = asof.strftime("%B %Y · As-of %Y-%m-%d")
if dash == "Rates":
    scores = score_rates(data, asof, weights)
    html = render_rates(scores, asof_str, weights)
else:
    scores = score_equity(data, asof, weights)
    html = render_equity(scores, asof_str, weights)

components.html(html, height=950, scrolling=True)

if show_raw:
    st.markdown("#### Raw factor table")
    st.dataframe(
        scores.style.format({c: "{:+.2f}" for c in scores.columns
                             if c not in ("incomplete", "country", "name", "region")}),
        use_container_width=True, height=460,
    )

if show_diag:
    st.markdown("#### Data coverage")
    diag = []
    for name, df in data.items():
        diag.append({
            "sheet": name,
            "rows": len(df),
            "cols": df.shape[1],
            "first": df.index.min().date() if len(df) else "n/a",
            "last":  df.index.max().date() if len(df) else "n/a",
            "missing_at_asof": int(df.loc[:asof].ffill().iloc[-1].isna().sum()) if len(df.loc[:asof]) else df.shape[1],
        })
    st.dataframe(pd.DataFrame(diag), use_container_width=True)
    incomp = scores[scores["incomplete"]].index.tolist()
    if incomp:
        st.warning(f"Incomplete coverage (orange-pipped): {', '.join(map(str, incomp))}")

col1, col2 = st.columns(2)
with col1:
    st.download_button(
        f"Download {dash.lower()} scores (CSV)",
        scores.to_csv().encode(),
        file_name=f"pulsar_{dash.lower()}_{asof.date()}.csv",
        mime="text/csv",
    )
with col2:
    st.download_button(
        f"Download {dash.lower()} dashboard (HTML)",
        html.encode(),
        file_name=f"pulsar_{dash.lower()}_{asof.date()}.html",
        mime="text/html",
    )
